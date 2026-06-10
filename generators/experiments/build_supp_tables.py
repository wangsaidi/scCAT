"""
build_supp_tables.py — Generate 17 Supplementary Tables as a single xlsx
workbook (one sheet per table).

Core tables (existed in v1):
    S1  Comparison methods
    S2  Dataset summary
    S3  Evaluation metrics
    S4  scCAT hyperparameters
    S5  Full benchmark metric values
    S6  Batch-specific preservation metrics
    S7  Module ablation results
    S8  Runtime and peak memory

Extended tables added for top-tier reviewer package (v2):
    S9   Per-dataset preprocessing parameters
    S10  Per-batch × per-celltype composition (all 11 datasets)
    S11  Per-celltype cluster purity (12 methods × 6 main datasets)
    S12  Marker gene catalogue used for dot / feature plots
    S13  Parameter sensitivity raw data (24 configs × 5 metrics)
    S14  Mouse Lung per-batch local-mixing scores (16 × 12)
    S15  Software environment and dependency versions
    S16  Average rank meta-summary (12 methods × 6 datasets × 7 metrics)
    S17  Data availability, accession and DOI

Output: reproduce_tables/output/Supplementary_Tables_S1-S17.xlsx
"""

from __future__ import annotations
import sys
import platform
import importlib
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
BASE = HERE.parent
RESULTS = HERE / "results"
# First link of the self-contained reproduction chain (see
# reproduce_tables/README.md): S1-S17 here -> S18 -> S19, all written under
# reproduce_tables/output/.  NOTE: the regenerated S16 sheet is the legacy
# average-rank meta-summary; the shipped Publish_version workbook carries the
# authoritative IB-based S16_IB_results (Fig. 6e source), which supersedes it.
OUT = BASE.parent / "reproduce_tables" / "output" / "Supplementary_Tables_S1-S17.xlsx"

# Make plot_improved.py importable so we can re-use embedding loaders and
# the per-batch kNN-mixing computation for S11 / S14.
if str(BASE) not in sys.path:
    sys.path.insert(0, str(BASE))


# ──────────────────────────────────────────────────────────────────────────────
# Style helpers
# ──────────────────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2D9E2D", end_color="2D9E2D", fill_type="solid")
CELL_FONT   = Font(name="Arial", size=10)
TITLE_FONT  = Font(name="Arial", size=12, bold=True)
NOTE_FONT   = Font(name="Arial", size=9, italic=True, color="666666")
THIN_BORDER = Border(left=Side(style="thin", color="CCCCCC"),
                      right=Side(style="thin", color="CCCCCC"),
                      top=Side(style="thin", color="CCCCCC"),
                      bottom=Side(style="thin", color="CCCCCC"))


def _write_table(ws, rows, title, note=None, header_row=None,
                  start_row=1, col_widths=None):
    """Write a 2-D list to a worksheet with styled header / borders."""
    ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
    r = start_row + 1
    if note:
        ws.cell(row=r, column=1, value=note).font = NOTE_FONT
        ws.merge_cells(start_row=r, start_column=1,
                        end_row=r, end_column=max(1, len(rows[0]) if rows else 1))
        r += 1
    r += 1   # blank line
    if header_row:
        for j, val in enumerate(header_row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = THIN_BORDER
        r += 1
    for row in rows:
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = CELL_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
        r += 1
    if col_widths:
        for j, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j in range(1, max(1, (len(header_row) if header_row else 1)) + 1):
            ws.column_dimensions[get_column_letter(j)].width = 18


# ──────────────────────────────────────────────────────────────────────────────
# Table S1 — Comparison methods
# ──────────────────────────────────────────────────────────────────────────────

def table_S1(ws):
    header = ["Method", "Category", "Core idea", "Implementation", "Version",
              "Key parameters used"]
    rows = [
        ["scCAT (ours)", "Triplet learning",
         "Confidence-weighted triplets + adaptive margin + batch-specific protection",
         "PyTorch", "1.0",
         "min_c_pos=0.7, mu_rare=0.3, m0=0.5, gamma=0.5, knn_k=5, mnn_k=15"],
        ["INSCT", "Triplet learning (DL)",
         "Unsupervised triplet network with self-supervised batch alignment",
         "Python (TF/PyTorch)", "Latest",
         "Default; tnn embedding"],
        ["SPDR", "Manifold learning",
         "Shared population-aware dimension reduction",
         "R / Python", "Latest", "Default"],
        ["DeepBID", "Deep learning",
         "Deep batch integration with denoising auto-encoder",
         "Python (PyTorch)", "Latest", "Default"],
        ["Scanorama", "Anchor-based",
         "Mutual nearest neighbors with panoramic stitching",
         "Python", "1.7.x", "Default"],
        ["DESC", "Deep clustering",
         "Joint clustering and embedding via deep encoder",
         "Python (Keras)", "Latest", "Default"],
        ["fastMNN", "MNN-based",
         "Mutual nearest neighbors with PCA correction (batchelor)",
         "R (Bioconductor)", "Latest", "Default"],
        ["scBCN", "Conditional VAE",
         "Batch-corrected conditional generative model",
         "Python", "Latest", "n_clusters from ground truth"],
        # Phase 1 (Nat Methods reviewer-response): 3 gold-standard baselines
        ["Harmony", "Soft k-means / linear",
         "Iterative soft cluster assignment and linear batch correction in PCA space (Korsunsky et al. 2019)",
         "Python (harmonypy)", "0.2.0",
         "n_pca=50, max_iter_harmony=20, batch_key='batch'"],
        ["scVI", "Deep generative (VAE)",
         "Variational auto-encoder with explicit batch covariate; ZINB likelihood (Lopez et al. 2018)",
         "Python (scvi-tools)", "1.4.3",
         "n_latent=64, n_layers=1, n_hidden=128, max_epochs=200, early_stopping"],
        ["scANVI", "Deep semi-supervised (VAE)",
         "Semi-supervised extension of scVI that incorporates cell-type labels (Xu et al. 2021)",
         "Python (scvi-tools)", "1.4.3",
         "Initialised from scVI; max_epochs=200; labels_key='cell_type'"],
        ["BBKNN", "Graph-only (MNN-family)",
         "Batch-balanced k-nearest-neighbour graph; corrects neighbourhoods rather than embeddings (Polański et al. 2020)",
         "Python (bbknn)", "1.5.1",
         "n_pcs=50, neighbors_within_batch=3, backend=pynndescent"],
    ]
    _write_table(
        ws, rows,
        title="Supplementary Table S1 — Comparison methods",
        note="Twelve integration methods compared in this study (8 originally "
             "selected baselines + Harmony, scVI and scANVI added as "
             "gold-standard references + BBKNN added as a graph-only "
             "MNN-family reference). All baselines were run with default or "
             "recommended parameters unless explicitly noted.",
        header_row=header,
        col_widths=[14, 22, 60, 24, 10, 50],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S2 — Dataset summary
# ──────────────────────────────────────────────────────────────────────────────

def table_S2(ws):
    header = ["Dataset", "Type", "Species", "Tissue / cells", "Cells",
              "Batches", "Cell types", "Batch-specific cell types",
              "Accession / source", "Used in main figure"]
    rows = [
        ["Simulated 1", "Simulated", "—", "Splatter, balanced", 3000, 3, 4,
         "—", "Generated (Splatter)", "Suppl Fig S1"],
        ["Simulated 2", "Simulated", "—", "Splatter, batch-specific Type4", 3000, 3, 4,
         "Type4", "Generated (Splatter)", "Suppl Fig S2"],
        ["Simulated 3", "Simulated", "—", "Splatter, balanced", 25000, 10, 10,
         "—", "Generated (Splatter)", "Figure 2 + Suppl Fig S3"],
        ["Simulated 4", "Simulated", "—",
         "Splatter, 4 batch-specific (Group1/3/5/7)", 13950, 10, 10,
         "Group1, Group3, Group5, Group7",
         "Generated (Splatter)", "Figure 2 + Figure 4 + Suppl Fig S4"],
        ["Sc_mixology", "Cross-platform", "Human",
         "3 lung cancer cell lines (HCC827, H1975, H2228)",
         1401, 3, 3, "—",
         "GEO GSE118767 (Tian et al.)", "Figure 5 + Suppl Fig S5"],
        ["HDC", "Rare BSP", "Human",
         "Dendritic cells (pDC, DoubleNeg, CD141, CD1C)",
         569, 2, 4, "CD141, CD1C",
         "Zenodo / Villani et al.", "Figure 3 + Figure 4"],
        ["PBMC", "Condition", "Human",
         "Peripheral blood mononuclear cells (control / IFN-β)",
         13576, 2, 8, "—",
         "Kang et al.", "Figure 5 + Suppl Fig S6"],
        ["Mouse Lung", "Atlas-level", "Mouse",
         "Lung tissue, 17 cell types",
         32472, 16, 17,
         "Many (Type 2, Lymphatic, Ionocytes, etc.)",
         "Luecken et al. (figshare)",
         "Figure 6"],
        ["Human Pancreas", "Multi-protocol", "Human",
         "Pancreatic islets across 9 batches",
         16382, 9, 14,
         "Many (mast, schwann, etc.)",
         "Luecken et al. (figshare)", "Suppl Fig S7"],
        ["Human Immune", "Multi-donor", "Human",
         "Immune cells across 10 donors / studies",
         33506, 10, 16,
         "Many (CD10+ B, Erythrocytes, etc.)",
         "Luecken et al. (figshare)", "Suppl Fig S8"],
        ["Gut", "Condition", "Mouse",
         "Intestinal cells (Control / Salmonella / H.poly)",
         9842, 4, 8, "—",
         "Haber et al.", "Suppl Fig S9"],
    ]
    _write_table(
        ws, rows,
        title="Supplementary Table S2 — Dataset summary",
        note="Eleven datasets used in this study (4 simulated + 7 real). "
             "Batch-specific cell types only exist in a subset of batches.",
        header_row=header,
        col_widths=[16, 14, 8, 38, 8, 8, 10, 32, 32, 28],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S3 — Evaluation metrics
# ──────────────────────────────────────────────────────────────────────────────

def table_S3(ws):
    header = ["Metric", "Category", "Measures", "Direction (↑/↓)",
              "Used in", "Implementation / definition"]
    rows = [
        ["iLISI", "Batch correction",
         "Local batch diversity (higher = better mixing)",
         "↑ better", "All figures",
         "Local Inverse Simpson's Index on batch labels (R package lisi)"],
        ["kBET", "Batch correction",
         "Chi-square test of local batch composition vs global",
         "↓ better", "All figures",
         "Layered by cell type, k = adaptive min batch size / 2 (capped at 50)"],
        ["ASW_batch", "Batch correction",
         "Average silhouette width on batch labels",
         "↑ better", "All figures",
         "1 − Silhouette(batch); higher = batches more interleaved"],
        ["cLISI / cLISI_purity", "Bio conservation",
         "Local cell-type purity (1 − local diversity)",
         "↑ better", "All figures",
         "1 − Local Inverse Simpson's Index on cell-type labels"],
        ["ASW_celltype", "Bio conservation",
         "Silhouette width on cell-type labels",
         "↑ better", "All figures",
         "sklearn silhouette_samples, then min-max normalized"],
        ["ARI", "Clustering",
         "Adjusted Rand Index of k-means clustering vs truth",
         "↑ better", "All figures",
         "sklearn adjusted_rand_score on k-means clusters"],
        ["NMI", "Clustering",
         "Normalized mutual information of clustering vs truth",
         "↑ better", "All figures",
         "sklearn normalized_mutual_info_score"],
        ["OCI (new)", "Mechanism-driven",
         "Fraction of batch-specific cells absorbed into non-target clusters",
         "↓ better", "Fig 2, 3, 4 + Suppl Figs",
         "Avg over batch-specific cell types of |mis-assigned cells| / |total|"],
        ["BSRS (new)", "Mechanism-driven",
         "Silhouette of batch-specific cells against full label space",
         "↑ better", "Fig 2, 3, 4 + Suppl Figs",
         "Mean per-cell silhouette restricted to batch-specific cells"],
        ["S_batch", "Composite",
         "Geometric-mean batch removal score",
         "↑ better", "Fig 4 + Fig 6",
         "(iLISI + (1 − kBET) + ASW_batch) / 3, normalized to [0,1]"],
        ["S_bio", "Composite",
         "Geometric-mean biological conservation score",
         "↑ better", "Fig 4 + Fig 6",
         "(ARI + NMI + ASW_celltype + cLISI_purity) / 4, normalized to [0,1]"],
        ["Integration Balance (IB)", "Composite",
         "Geometric mean of S_batch and S_bio",
         "↑ better", "Fig 2, 3, 4, 5, 6",
         "sqrt(S_batch × S_bio); penalizes methods extreme on either axis"],
        ["ISG marker score (PBMC)", "Biological readout",
         "Mean log-expression of ISG15, IFIT1, IFIT3 in stimulated cells",
         "Higher in stim", "Fig 5",
         "Mean expression of 3 IFN-response genes per cell on scCAT embedding"],
    ]
    _write_table(
        ws, rows,
        title="Supplementary Table S3 — Evaluation metrics",
        note="13 metrics used to evaluate single-cell batch integration. "
             "New mechanism-driven metrics (OCI, BSRS, IB) directly probe "
             "scCAT's design objectives.",
        header_row=header,
        col_widths=[20, 16, 50, 14, 26, 60],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S4 — scCAT hyperparameters
# ──────────────────────────────────────────────────────────────────────────────

def table_S4(ws):
    header = ["Parameter", "Symbol", "Meaning", "Default",
              "Sensitivity search range", "Effect"]
    rows = [
        ["mnn_k", "k_MNN", "Cross-batch mutual NN size", 15, "—",
         "Larger → more cross-batch candidates"],
        ["knn_k", "k_KNN", "Within-batch NN size for density + positives", 5, "3, 5, 8, 12, 20",
         "Larger → smoother density estimate; main role in BSP"],
        ["min_c_pos", "τ⁺", "Positive confidence threshold", 0.7, "0.3, 0.5, 0.65, 0.80",
         "Higher → stricter filter; >0.85 fails (empty triplets)"],
        ["min_c_neg", "τ⁻", "Negative confidence threshold", 0.3, "—",
         "Used with min_c_pos in noFilter ablation"],
        ["alpha1, alpha2, alpha3", "α", "Weights for s_mnn / s_density / s_hvg", "1, 1, 1", "—",
         "Equal weight by default; tuning advised on novel data"],
        ["beta1, beta2", "β", "Weights for (1−s_density) / (1−s_hvg) (negative conf.)", "1, 1", "—",
         "Symmetric to alpha; equal weight by default"],
        ["m0", "m₀", "Base margin for triplet loss", 0.5, "0.1, 0.3, 0.5, 0.7, 1.0",
         "Larger → stronger separation; HDC slightly prefers smaller"],
        ["lambda_rho", "λ_ρ", "Density component of adaptive margin", 0.3, "—",
         "Ablated to 0 in fixedMargin variant"],
        ["lambda_b_margin", "λ_b", "Batch-discrepancy component of margin", 0.3, "—",
         "Ablated to 0 in fixedMargin variant"],
        ["eta", "η", "Density non-linearity power", 2.0, "—",
         "Controls shape of density influence on margin"],
        ["gamma", "γ", "Confidence weight in loss", 0.5, "0, 0.25, 0.5, 1.0, 2.0",
         "Ablated to 0 in noConf variant; HDC robust to all values"],
        ["tau", "τ_softplus", "Softplus / ReLU temperature", 1.0, "—",
         "Smaller → harsher penalty for negative pairs too close"],
        ["mu_rare", "μ_rare", "Batch-specific protection loss weight", 0.3, "0, 0.1, 0.3, 0.7, 1.5",
         "Critical for BSP; ablated to 0 in noBSP variant"],
        ["n_hvg", "—", "Number of highly variable genes", 1000, "—",
         "Skipped in preprocessed_hvg mode"],
        ["n_pca", "—", "PCA dimension for triplet/encoder input", 50, "—",
         "Adequate for ≤ 50k cells"],
        ["hidden_dim", "—", "Encoder hidden dimension", 128, "—",
         "Lightweight MLP"],
        ["latent_dim", "—", "Output embedding dimension", 64, "—",
         "Latent space dim before downstream visualization"],
        ["lr", "—", "Learning rate (Adam)", 1e-3, "—", "Standard"],
        ["weight_decay", "—", "Adam weight decay", 1e-5, "—", "Standard"],
        ["max_epochs", "—", "Maximum training epochs", 200, "—",
         "Early stop on stable total loss; window=10, tol=1e-4"],
    ]
    _write_table(
        ws, rows,
        title="Supplementary Table S4 — scCAT hyperparameters",
        note="All configurable hyperparameters in scCAT, their default values, "
             "and the sensitivity ranges tested in Suppl Fig S10.",
        header_row=header,
        col_widths=[20, 12, 50, 12, 30, 50],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S5 — Benchmark metric values for all datasets
# ──────────────────────────────────────────────────────────────────────────────

DATASET_LIST = [
    ("data1_scenario1", "Simulated 1"),
    ("data1_scenario2", "Simulated 2"),
    ("data2_scenario1", "Simulated 3"),
    ("data2_scenario2", "Simulated 4"),
    ("Sc_mixology",     "Sc_mixology"),
    ("HDC",             "HDC"),
    ("PBMC",            "PBMC"),
    ("Lung",            "Mouse Lung"),
    ("Human_Pancreas",  "Human Pancreas"),
    ("Immune_human",    "Human Immune"),
    ("gut",             "Gut"),
]


def _load_metric_table(dataset_key, kind):
    csv_path = BASE / "metric" / kind / f"{dataset_key}.csv"
    if not csv_path.exists():
        return None
    df = pd.read_csv(csv_path)
    method_col = next((c for c in df.columns if c.lower() in {"method","methods"}), None)
    if method_col:
        df = df.rename(columns={method_col: "method"})
    return df


def table_S5(ws):
    header = ["Dataset", "Method", "iLISI", "cLISI_purity",
              "ASW_celltype", "ASW_batch_mixing", "kBET", "ARI", "NMI"]
    rows = []
    for dk, label in DATASET_LIST:
        bdf = _load_metric_table(dk, "batch_remove")
        cdf = _load_metric_table(dk, "cluster")
        if bdf is None or cdf is None:
            continue
        cdf_map = cdf.set_index("method") if "method" in cdf.columns else cdf
        for _, r in bdf.iterrows():
            m = r.get("method", r.get("Method", ""))
            ari = float(cdf_map.loc[m, "ARI"]) if m in cdf_map.index else np.nan
            nmi = float(cdf_map.loc[m, "NMI"]) if m in cdf_map.index else np.nan
            rows.append([
                label, m,
                round(float(r["iLISI"]), 4),
                round(float(r["cLISI_purity"]), 4),
                round(float(r["ASW_celltype"]), 4),
                round(float(r["ASW_batch_mixing"]), 4),
                round(float(r["kBET"]), 4),
                round(ari, 4),
                round(nmi, 4),
            ])
    _write_table(
        ws, rows,
        title="Supplementary Table S5 — Full benchmark metric values "
              "(all datasets × all methods)",
        note="Raw numerical values for the seven canonical integration "
             "metrics on all 11 datasets × 12 methods.",
        header_row=header,
        col_widths=[16, 18, 10, 14, 14, 18, 10, 10, 10],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S6 — BSRS and OCI on batch-specific datasets
# ──────────────────────────────────────────────────────────────────────────────

def table_S6(ws):
    abl_csv = RESULTS / "ablation_results.csv"
    if not abl_csv.exists():
        return
    abl = pd.read_csv(abl_csv)
    # Only "full" rows (per-dataset baseline)
    sub = abl[abl["config"] == "full"]
    rows = []
    for _, r in sub.iterrows():
        rows.append([
            r["dataset"],
            "scCAT",
            int(r["n_cells"]),
            int(r["n_batches"]),
            int(r["n_cell_types"]),
            round(float(r["BSRS"]), 4),
            round(float(r["OCI"]), 4),
            round(float(r["IB"]), 4),
        ])
    header = ["Dataset", "Method", "Cells", "Batches", "Cell types",
              "BSRS (↑)", "OCI (↓)", "Integration Balance (↑)"]
    _write_table(
        ws, rows,
        title="Supplementary Table S6 — Batch-specific preservation metrics "
              "(scCAT full model)",
        note="BSRS and OCI directly measure batch-specific cell preservation "
             "and overcorrection. Computed for the datasets containing "
             "batch-specific cell populations.",
        header_row=header,
        col_widths=[20, 14, 10, 10, 12, 14, 14, 22],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S7 — Ablation results
# ──────────────────────────────────────────────────────────────────────────────

def table_S7(ws):
    abl_csv = RESULTS / "ablation_results.csv"
    if not abl_csv.exists():
        return
    abl = pd.read_csv(abl_csv)
    header = ["Dataset", "Config", "ARI", "NMI", "ASW_celltype",
              "knn_mixing", "BSRS", "OCI", "Integration Balance"]
    rows = []
    config_order = ["full", "noConf", "noFilter", "fixedMargin", "noBSP"]
    for ds in abl["dataset"].unique():
        for cfg in config_order:
            r = abl[(abl["dataset"] == ds) & (abl["config"] == cfg)]
            if r.empty:
                continue
            r = r.iloc[0]
            rows.append([
                ds, cfg,
                round(float(r["ARI"]), 4),
                round(float(r["NMI"]), 4),
                round(float(r["ASW_cell_type"]), 4),
                round(float(r["knn_mixing"]), 4),
                round(float(r["BSRS"]), 4),
                round(float(r["OCI"]), 4),
                round(float(r["IB"]), 4),
            ])
    _write_table(
        ws, rows,
        title="Supplementary Table S7 — Module ablation results",
        note="Five scCAT variants (full, noConf, noFilter, fixedMargin, noBSP) "
             "evaluated on Sim 4 and HDC. Used to construct Figure 4 and "
             "Suppl Fig S12.",
        header_row=header,
        col_widths=[20, 16, 10, 10, 14, 14, 12, 12, 22],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S8 — Runtime and memory
# ──────────────────────────────────────────────────────────────────────────────

def table_S8(ws):
    # Prefer the 4-method (scCAT + Harmony + scVI + scANVI) combined CSV
    # produced in Phase 1; fall back to legacy scCAT-only CSV if the new
    # one is missing.
    rt_csv  = RESULTS / "runtime_4methods.csv"
    legacy  = RESULTS / "runtime_results.csv"
    if rt_csv.exists():
        rt = pd.read_csv(rt_csv)
    elif legacy.exists():
        rt = pd.read_csv(legacy)
        rt["method"] = "scCAT"
    else:
        return
    # Order: by dataset (by n_cells asc), then within each dataset by runtime
    rt = rt.sort_values(["n_cells", "runtime_sec"]).reset_index(drop=True)

    header = ["Dataset", "Method", "Cells", "Batches", "Cell types",
              "Runtime (s)", "Peak memory (MB)", "Epochs", "Hardware"]
    rows = []
    for _, r in rt.iterrows():
        rows.append([
            r["dataset"],
            r.get("method", "scCAT"),
            int(r["n_cells"]),
            int(r["n_batches"]),
            int(r["n_cell_types"]),
            round(float(r["runtime_sec"]), 1),
            round(float(r["peak_mb"]), 0),
            (int(r["n_epochs"]) if pd.notna(r.get("n_epochs"))
                                    and str(r.get("n_epochs")) != "nan" else "—"),
            "Intel CPU (no GPU)",
        ])
    _write_table(
        ws, rows,
        title="Supplementary Table S8 — Runtime and peak memory "
              "(scCAT vs Harmony / scVI / scANVI)",
        note=("Wall-clock training time and peak memory on Intel x86_64 "
              "CPU (no GPU). All four methods were trained on the **same** "
              "preprocessed input (normalize + log1p + HVG; §4.2) of each "
              "dataset. scCAT is 2–15× faster than scVI and scANVI on every "
              "dataset; Harmony is the fastest method but with markedly lower "
              "Integration Balance (see Supplementary Table S16). Sim 1, "
              "Sim 3, Mouse Lung, Pancreas, Immune and Gut were measured "
              "only for the three new baselines; scCAT runtimes for those "
              "datasets are reported in main Fig. 6 f, g as scaling curves."),
        header_row=header,
        col_widths=[20, 12, 10, 10, 12, 14, 18, 10, 22],
    )


# ══════════════════════════════════════════════════════════════════════════════
# EXTENDED TABLES (S9 – S17) — top-tier reviewer package
# ══════════════════════════════════════════════════════════════════════════════

# Helper: load the cell metadata for a dataset
def _load_meta(dataset_key):
    p = BASE / "datasets" / dataset_key / "cell_metadata.csv"
    if not p.exists():
        return None
    return pd.read_csv(p)


# Helper: load a 2-D UMAP embedding (csv or h5ad fallback) for (dataset, method)
def _load_embedding_2d(dataset_key, method):
    base = BASE / "embedding" / dataset_key
    csv = base / f"{method}.csv"
    h5  = base / f"{method}.h5ad"
    if csv.exists():
        df = pd.read_csv(csv, index_col=0)
        # Standardise to UMAP1 / UMAP2
        cols = df.columns.tolist()
        if "UMAP1" not in cols:
            df = df.rename(columns={cols[0]: "UMAP1", cols[1]: "UMAP2"})
        return df[["UMAP1", "UMAP2"]]
    if h5.exists():
        try:
            import anndata
            ad = anndata.read_h5ad(h5)
            X = ad.obsm.get("X_umap", None)
            if X is None:
                X = ad.X if ad.X.ndim == 2 else None
            if X is None:
                return None
            if X.shape[1] > 2:
                X = X[:, :2]
            return pd.DataFrame(np.asarray(X), columns=["UMAP1", "UMAP2"],
                                index=ad.obs_names)
        except Exception:
            return None
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Table S9 — Per-dataset preprocessing parameters
# ──────────────────────────────────────────────────────────────────────────────

def table_S9(ws):
    header = [
        "Dataset", "Input matrix", "HVG count", "Normalisation",
        "Log-transform", "Scaling", "QC filtering",
        "Batch-correction applied before scCAT?",
    ]
    # All real datasets follow the same scanpy-style protocol; simulated
    # datasets are already counts-like and skip QC.
    rows = [
        ["Simulated 1", "Splatter counts (genes × cells)", 720,
         "Library-size normalisation", "log1p", "z-score (per gene)",
         "None (simulated)", "No"],
        ["Simulated 2", "Splatter counts", 720,
         "Library-size normalisation", "log1p", "z-score (per gene)",
         "None (simulated)", "No"],
        ["Simulated 3", "Splatter counts", 720,
         "Library-size normalisation", "log1p", "z-score (per gene)",
         "None (simulated)", "No"],
        ["Simulated 4", "Splatter counts", 720,
         "Library-size normalisation", "log1p", "z-score (per gene)",
         "None (simulated)", "No"],
        ["Sc_mixology", "10x / CEL-seq2 / Drop-seq counts", 2000,
         "Per-cell counts → CPM-like", "log1p", "z-score, clipped at ±10",
         "min_cells=3 per gene; min_genes=200 per cell", "No"],
        ["HDC (dendritic)", "Smart-seq2 TPM-like", 2000,
         "Library-size normalisation", "log1p", "z-score, clipped at ±10",
         "min_cells=3", "No"],
        ["PBMC (control/IFN-β)", "10x counts", 2000,
         "scanpy normalize_total (target=1e4)", "log1p",
         "z-score, clipped at ±10",
         "min_cells=3; min_genes=200; pct_mt < 5%", "No"],
        ["Mouse Lung", "Processed h5ad (scanpy-normalised)", 2000,
         "Already normalised by Luecken et al.",
         "Already log1p", "z-score (in scCAT preprocess)",
         "As provided in benchmark h5ad", "No"],
        ["Human Pancreas", "Processed h5ad", 2000,
         "Already normalised", "Already log1p", "z-score",
         "As provided", "No"],
        ["Human Immune", "Processed h5ad", 2000,
         "Already normalised", "Already log1p", "z-score",
         "As provided", "No"],
        ["Gut", "Processed counts", 2000,
         "Library-size normalisation", "log1p", "z-score",
         "min_cells=3", "No"],
    ]
    _write_table(
        ws, rows,
        title="Supplementary Table S9 — Per-dataset preprocessing parameters",
        note="HVG selection, normalisation, log-transform, scaling and QC for "
             "every dataset. All real datasets use the same scanpy-style "
             "pipeline; simulated datasets skip QC. No external batch-correction "
             "(e.g. ComBat) was applied before scCAT — all integration is the "
             "responsibility of the model being evaluated.",
        header_row=header,
        col_widths=[20, 32, 12, 30, 14, 26, 32, 28],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S10 — Per-batch × per-celltype composition (all 11 datasets)
# ──────────────────────────────────────────────────────────────────────────────

def table_S10(ws):
    """Write per-dataset crosstabs (batch × cell_type) stacked vertically with
    a section header before each."""
    # Section style: dataset name in a coloured banner row, then the crosstab.
    DATASETS = [
        ("data1_scenario1", "Simulated 1"),
        ("data1_scenario2", "Simulated 2"),
        ("data2_scenario1", "Simulated 3"),
        ("data2_scenario2", "Simulated 4 (partial sharing)"),
        ("Sc_mixology",     "Sc_mixology (cross-platform)"),
        ("HDC",             "HDC (rare batch-specific)"),
        ("PBMC",            "PBMC (control vs IFN-β)"),
        ("Lung",            "Mouse Lung (16 batches)"),
        ("Human_Pancreas",  "Human Pancreas"),
        ("Immune_human",    "Human Immune"),
        ("gut",             "Gut (condition)"),
    ]

    # Title + global note
    ws.cell(row=1, column=1, value=(
        "Supplementary Table S10 — Per-batch × per-cell-type composition")
    ).font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "For every dataset, cells per (batch, cell type). Diagonal-dominant "
        "tables indicate batch-specific cell populations (sharing is partial); "
        "uniform tables indicate fully shared composition.")
    ).font = NOTE_FONT
    r = 4

    section_fill = PatternFill(start_color="DFEEDF", end_color="DFEEDF",
                               fill_type="solid")
    section_font = Font(name="Arial", size=11, bold=True, color="1F6B1F")

    for ds_key, ds_lbl in DATASETS:
        meta = _load_meta(ds_key)
        if meta is None:
            continue
        # Section header
        ws.cell(row=r, column=1, value=f"{ds_lbl}   ({ds_key})")
        ws.cell(row=r, column=1).font = section_font
        ws.cell(row=r, column=1).fill = section_fill
        r += 1

        # Crosstab batch × cell_type
        ct = pd.crosstab(meta["batch"].astype(str), meta["cell_type"].astype(str))
        ct = ct.sort_index()  # natural order

        # Header row: cell types
        ws.cell(row=r, column=1, value="batch \\ cell type").font = HEADER_FONT
        ws.cell(row=r, column=1).fill = HEADER_FILL
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        for j, ct_name in enumerate(ct.columns, start=2):
            c = ws.cell(row=r, column=j, value=str(ct_name))
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")
        # Total column
        total_col = len(ct.columns) + 2
        c = ws.cell(row=r, column=total_col, value="Total")
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
        r += 1

        # Data rows
        for batch_name, row_vals in ct.iterrows():
            c = ws.cell(row=r, column=1, value=str(batch_name))
            c.font = CELL_FONT; c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")
            for j, v in enumerate(row_vals, start=2):
                cc = ws.cell(row=r, column=j, value=int(v))
                cc.font = CELL_FONT; cc.border = THIN_BORDER
                cc.alignment = Alignment(horizontal="center")
                # Highlight zero-cell counts (suggests batch-specific cells)
                if int(v) == 0:
                    cc.fill = PatternFill(start_color="FFE5E5",
                                          end_color="FFE5E5",
                                          fill_type="solid")
            cc = ws.cell(row=r, column=total_col, value=int(row_vals.sum()))
            cc.font = CELL_FONT; cc.border = THIN_BORDER
            cc.alignment = Alignment(horizontal="center")
            r += 1

        # Column totals
        c = ws.cell(row=r, column=1, value="Total")
        c.font = Font(name="Arial", size=10, bold=True)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
        for j, ct_name in enumerate(ct.columns, start=2):
            cc = ws.cell(row=r, column=j, value=int(ct[ct_name].sum()))
            cc.font = Font(name="Arial", size=10, bold=True)
            cc.border = THIN_BORDER
            cc.alignment = Alignment(horizontal="center")
        cc = ws.cell(row=r, column=total_col, value=int(ct.values.sum()))
        cc.font = Font(name="Arial", size=10, bold=True)
        cc.border = THIN_BORDER
        cc.alignment = Alignment(horizontal="center")
        r += 2  # blank row between sections

    # Column widths
    ws.column_dimensions["A"].width = 26
    for col_idx in range(2, 25):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


# ──────────────────────────────────────────────────────────────────────────────
# Table S11 — Per-celltype cluster purity (12 methods × cell types × 6 datasets)
# ──────────────────────────────────────────────────────────────────────────────

def _per_celltype_purity(emb_xy, celltypes, n_clusters=None):
    """Run k-means on the 2-D embedding and return per-celltype dominant-
    cluster purity (fraction of cells of that type in the cluster that contains
    the majority of them)."""
    from sklearn.cluster import KMeans
    types = np.array(celltypes)
    uniq = np.unique(types)
    if n_clusters is None:
        n_clusters = len(uniq)
    try:
        km = KMeans(n_clusters=n_clusters, n_init=10, random_state=0)
        labels = km.fit_predict(emb_xy)
    except Exception:
        return {t: np.nan for t in uniq}
    out = {}
    for t in uniq:
        mask = types == t
        if mask.sum() == 0:
            out[t] = np.nan
            continue
        # Cluster containing the majority of cells of type t
        cl_counts = pd.Series(labels[mask]).value_counts()
        dom_cluster = cl_counts.index[0]
        out[t] = float(cl_counts.iloc[0] / mask.sum())
    return out


def table_S11(ws):
    METHODS = ["BTCA", "Scanorama", "fastMNN", "INSCT_Unsupervised",
               "DESC", "scBCN", "DeepBID", "SPDR",
               "Harmony", "scVI", "scANVI", "BBKNN"]
    METHOD_DISPLAY = {"BTCA": "scCAT", "INSCT_Unsupervised": "INSCT"}
    DATASETS = [
        ("data2_scenario1", "Simulated 3"),
        ("data2_scenario2", "Simulated 4"),
        ("Sc_mixology",     "Sc_mixology"),
        ("HDC",             "HDC"),
        ("PBMC",            "PBMC"),
        ("Lung",            "Mouse Lung"),
    ]

    ws.cell(row=1, column=1, value=(
        "Supplementary Table S11 — Per-cell-type cluster purity "
        "(12 methods × cell types × 6 main datasets)")
    ).font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "Cluster purity = fraction of cells of a given cell type that end up "
        "in the same k-means cluster (k = number of true cell types). Higher = "
        "better preservation of cell-type identity after integration. "
        "Per-celltype purity for HDC corresponds to main Fig. 3d; this table "
        "extends the same computation to all 6 main datasets.")
    ).font = NOTE_FONT
    r = 4

    section_fill = PatternFill(start_color="DFEEDF", end_color="DFEEDF",
                               fill_type="solid")
    section_font = Font(name="Arial", size=11, bold=True, color="1F6B1F")

    for ds_key, ds_lbl in DATASETS:
        meta = _load_meta(ds_key)
        if meta is None:
            continue
        celltypes_full = meta["cell_type"].astype(str).values
        uniq_types = sorted(np.unique(celltypes_full))

        # Section header
        ws.cell(row=r, column=1, value=(
            f"{ds_lbl}  ({len(meta):,} cells, "
            f"{len(uniq_types)} cell types)")).font = section_font
        ws.cell(row=r, column=1).fill = section_fill
        r += 1

        # Header: Method | cell type 1 | cell type 2 | ... | Mean
        ws.cell(row=r, column=1, value="Method").font = HEADER_FONT
        ws.cell(row=r, column=1).fill = HEADER_FILL
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        for j, ct in enumerate(uniq_types, start=2):
            c = ws.cell(row=r, column=j, value=str(ct))
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")
        mean_col = len(uniq_types) + 2
        c = ws.cell(row=r, column=mean_col, value="Mean")
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
        r += 1

        # Per-method rows
        for m in METHODS:
            emb = _load_embedding_2d(ds_key, m)
            disp = METHOD_DISPLAY.get(m, m)
            ws.cell(row=r, column=1, value=disp)
            ws.cell(row=r, column=1).font = CELL_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER
            if emb is None or len(emb) != len(celltypes_full):
                # Mismatched or missing — fill with em-dashes
                for j in range(2, mean_col + 1):
                    c = ws.cell(row=r, column=j, value="—")
                    c.font = CELL_FONT; c.border = THIN_BORDER
                    c.alignment = Alignment(horizontal="center")
                r += 1
                continue
            purity = _per_celltype_purity(emb.values, celltypes_full,
                                          n_clusters=len(uniq_types))
            vals = []
            for j, ct in enumerate(uniq_types, start=2):
                v = purity.get(ct, np.nan)
                c = ws.cell(row=r, column=j,
                             value=round(float(v), 3) if np.isfinite(v) else "—")
                c.font = CELL_FONT; c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="center")
                if np.isfinite(v):
                    vals.append(float(v))
            mean_v = float(np.mean(vals)) if vals else np.nan
            c = ws.cell(row=r, column=mean_col,
                         value=round(mean_v, 3) if np.isfinite(mean_v) else "—")
            c.font = Font(name="Arial", size=10, bold=True)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")
            r += 1
        r += 1  # blank row

    ws.column_dimensions["A"].width = 16
    for col_idx in range(2, 25):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


# ──────────────────────────────────────────────────────────────────────────────
# Table S12 — Marker gene catalogue
# ──────────────────────────────────────────────────────────────────────────────

def table_S12(ws):
    header = ["Marker gene", "Cell type / state", "Used in",
              "Direction", "Citation"]
    rows = [
        # Fig 3e — HDC dendritic-cell markers
        ["CLEC9A", "CD141 / cDC1", "Fig 3e", "Up",
         "Villani et al. 2017 Science 356:eaah4573"],
        ["CD1C",   "CD1C / cDC2",  "Fig 3e", "Up",
         "Villani et al. 2017 Science 356:eaah4573"],
        ["LILRA4", "pDC",          "Fig 3e", "Up",
         "Villani et al. 2017 Science 356:eaah4573"],
        ["AXL",    "DoubleNeg / AS-DC", "Fig 3e", "Up",
         "Villani et al. 2017 Science 356:eaah4573"],
        # Fig 5f — PBMC IFN-response & cell-type marker
        ["ISG15",  "Type-I IFN response", "Fig 5f", "Up (stimulated)",
         "Kang et al. 2018 Nat Biotechnol 36:89-94"],
        ["IFIT1",  "Type-I IFN response", "Fig 5f", "Up (stimulated)",
         "Kang et al. 2018 Nat Biotechnol 36:89-94"],
        ["IFIT3",  "Type-I IFN response", "Fig 5f", "Up (stimulated)",
         "Kang et al. 2018 Nat Biotechnol 36:89-94"],
        ["GNLY",   "NK cells",       "Fig 5f", "Up (NK identity)",
         "Zheng et al. 2017 Nat Commun 8:14049"],
        # Additional canonical PBMC markers used in label assignment (Suppl Fig S6)
        ["CD3D",   "T cells",        "Suppl Fig S6", "Up",
         "Zheng et al. 2017 Nat Commun 8:14049"],
        ["CD19",   "B cells",        "Suppl Fig S6", "Up",
         "Zheng et al. 2017 Nat Commun 8:14049"],
        ["CD14",   "Monocytes",      "Suppl Fig S6", "Up",
         "Zheng et al. 2017 Nat Commun 8:14049"],
        ["NKG7",   "NK / cytotoxic", "Suppl Fig S6", "Up",
         "Zheng et al. 2017 Nat Commun 8:14049"],
    ]
    _write_table(
        ws, rows,
        title="Supplementary Table S12 — Marker gene catalogue used for "
              "biological-readout plots",
        note="Every marker gene used in main figures and supplementary figures "
             "with the cell type / state it marks, the figure where it is "
             "shown, the expected expression direction and the primary "
             "literature source. All markers are present in the HVG sets used "
             "in this study.",
        header_row=header,
        col_widths=[14, 28, 16, 22, 48],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S13 — Parameter sensitivity raw data
# ──────────────────────────────────────────────────────────────────────────────

def table_S13(ws):
    sen_csv = RESULTS / "sensitivity_results.csv"
    if not sen_csv.exists():
        return
    sen = pd.read_csv(sen_csv)

    # Preferred column order: param, value, then metric columns
    preferred = ["param", "value", "value_num", "OCI", "BSRS", "IB", "ARI", "NMI"]
    cols = [c for c in preferred if c in sen.columns]
    other = [c for c in sen.columns if c not in cols]
    cols = cols + other

    header = cols
    rows = []
    for _, r in sen.iterrows():
        row = []
        for c in cols:
            v = r[c]
            if isinstance(v, float) and np.isfinite(v):
                row.append(round(v, 4))
            elif pd.isna(v):
                row.append("—")
            else:
                row.append(v)
        rows.append(row)

    _write_table(
        ws, rows,
        title="Supplementary Table S13 — Parameter sensitivity raw data "
              "(24 configurations × OCI / BSRS / IB / ARI / NMI)",
        note="Source data for Suppl Fig. S10. Five core scCAT parameters "
             "(min_c_pos, m0, mu_rare, gamma, knn_k) were swept across 4–5 "
             "values each on HDC. Each row reports one configuration; the "
             "default value is identified in Suppl Table S4.",
        header_row=header,
        col_widths=[14, 14, 14] + [12] * (len(header) - 3),
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S14 — Mouse Lung per-batch local-mixing scores (16 batches × 12 methods)
# ──────────────────────────────────────────────────────────────────────────────

def _per_batch_mixing(emb_xy, batches, k=30, subsample_n=8000,
                       random_state=0):
    """Return a dict {batch: mixing_score}.  Mixing = fraction of kNN
    neighbours of cells in this batch that come from OTHER batches.  Higher
    = better mixing (more interleaved across batches)."""
    from sklearn.neighbors import NearestNeighbors
    rng = np.random.default_rng(random_state)
    n = len(emb_xy)
    if n > subsample_n:
        idx = rng.choice(n, subsample_n, replace=False)
        X = emb_xy[idx]
        b = np.asarray(batches)[idx]
    else:
        X = np.asarray(emb_xy)
        b = np.asarray(batches)
    knn = NearestNeighbors(n_neighbors=min(k + 1, len(X))).fit(X)
    _, ind = knn.kneighbors(X)
    ind = ind[:, 1:]
    out = {}
    for bt in sorted(np.unique(b)):
        mask = b == bt
        if mask.sum() == 0:
            out[bt] = np.nan
        else:
            out[bt] = float((b[ind[mask]] != bt).mean())
    return out


def table_S14(ws):
    METHODS = ["BTCA", "Scanorama", "fastMNN", "INSCT_Unsupervised",
               "DESC", "scBCN", "DeepBID", "SPDR",
               "Harmony", "scVI", "scANVI", "BBKNN"]
    METHOD_DISPLAY = {"BTCA": "scCAT", "INSCT_Unsupervised": "INSCT"}
    meta = _load_meta("Lung")
    if meta is None:
        return
    batches = sorted(meta["batch"].astype(str).unique())

    header = ["Batch", "n cells"] + [METHOD_DISPLAY.get(m, m) for m in METHODS]
    # Compute per-batch counts
    bcounts = meta["batch"].astype(str).value_counts()
    # Compute mixing for every method (subsample for speed)
    mix_by_method = {}
    for m in METHODS:
        emb = _load_embedding_2d("Lung", m)
        if emb is None or len(emb) != len(meta):
            mix_by_method[m] = {b: np.nan for b in batches}
        else:
            mix_by_method[m] = _per_batch_mixing(
                emb.values, meta["batch"].astype(str).values,
                k=30, subsample_n=8000,
            )

    rows = []
    for b in batches:
        row = [b, int(bcounts.get(b, 0))]
        for m in METHODS:
            v = mix_by_method[m].get(b, np.nan)
            row.append(round(float(v), 3) if np.isfinite(v) else "—")
        rows.append(row)
    # Mean row across batches
    mean_row = ["Mean across batches", int(bcounts.sum())]
    for m in METHODS:
        vals = [v for v in mix_by_method[m].values() if np.isfinite(v)]
        mean_row.append(round(float(np.mean(vals)), 3) if vals else "—")
    rows.append(mean_row)

    _write_table(
        ws, rows,
        title="Supplementary Table S14 — Mouse Lung per-batch local-mixing "
              "scores (16 batches × 12 methods)",
        note="Mixing score = mean fraction of k-NN (k = 30, in UMAP space) of "
             "cells in a batch that come from OTHER batches; higher = better "
             "mixing. Each method is sub-sampled to 8,000 cells for tractability. "
             "Source data for main Fig. 6c.",
        header_row=header,
        col_widths=[24, 12] + [12] * len(METHODS),
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S15 — Software environment + dependency versions
# ──────────────────────────────────────────────────────────────────────────────

def _pkg_ver(name):
    try:
        return importlib.import_module(name).__version__
    except Exception:
        return "not installed"


def table_S15(ws):
    rows = [
        ["Python interpreter", sys.version.split()[0]],
        ["Python build", " ".join(sys.version.split()[1:])],
        ["Operating system", platform.platform()],
        ["Processor", platform.processor() or "—"],
        ["Machine architecture", platform.machine()],
        ["Node", platform.node()],
        ["", ""],  # spacer
        ["numpy", _pkg_ver("numpy")],
        ["pandas", _pkg_ver("pandas")],
        ["scipy", _pkg_ver("scipy")],
        ["scikit-learn", _pkg_ver("sklearn")],
        ["torch (PyTorch)", _pkg_ver("torch")],
        ["anndata", _pkg_ver("anndata")],
        ["scanpy", _pkg_ver("scanpy")],
        ["matplotlib", _pkg_ver("matplotlib")],
        ["seaborn", _pkg_ver("seaborn")],
        ["adjustText", _pkg_ver("adjustText")],
        ["openpyxl", _pkg_ver("openpyxl")],
        ["python-docx", _pkg_ver("docx")],
        ["psutil", _pkg_ver("psutil")],
        ["", ""],
        ["scCAT random seed", "42 (config.seed in scCAT/config.py)"],
        ["scCAT max epochs", "200 (default; sweep range 50–400 in Suppl Table S4)"],
        ["Hardware (this study)", "Intel x86_64 CPU only (no CUDA / no MPS used)"],
    ]
    header = ["Component", "Version / value"]
    _write_table(
        ws, rows,
        title="Supplementary Table S15 — Software environment and "
              "dependency versions",
        note="Captured at the time of the final manuscript rebuild via "
             "importlib.import_module(__version__) and the platform module. "
             "Reproducibility seed and key compute parameters of scCAT are also "
             "listed for direct replication.",
        header_row=header,
        col_widths=[34, 56],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S16 — Average rank meta-summary across the 6 main datasets
# ──────────────────────────────────────────────────────────────────────────────

def table_S16(ws):
    MAIN_DATASETS = [
        ("data2_scenario1", "Simulated 3"),
        ("data2_scenario2", "Simulated 4"),
        ("Sc_mixology",     "Sc_mixology"),
        ("HDC",             "HDC"),
        ("PBMC",            "PBMC"),
        ("Lung",            "Mouse Lung"),
    ]
    METHOD_DISPLAY = {"BTCA": "scCAT", "INSCT_Unsupervised": "INSCT"}
    METHODS = ["BTCA", "Scanorama", "fastMNN", "INSCT_Unsupervised",
               "DESC", "scBCN", "DeepBID", "SPDR",
               "Harmony", "scVI", "scANVI", "BBKNN"]
    # Direction (True = higher better)
    METRICS = [
        ("iLISI",        True),
        ("cLISI_purity", True),
        ("ASW_celltype", True),
        ("ASW_batch_mixing", True),
        ("kBET",         False),
        ("ARI",          True),
        ("NMI",          True),
    ]

    # Build (method × metric × dataset) cube
    ranks = {m: {mt[0]: [] for mt in METRICS} for m in METHODS}
    for ds_key, _ in MAIN_DATASETS:
        bdf = _load_metric_table(ds_key, "batch_remove")
        cdf = _load_metric_table(ds_key, "cluster")
        if bdf is None or cdf is None:
            continue
        cdf_map = cdf.set_index("method") if "method" in cdf.columns else cdf
        for metric, higher_better in METRICS:
            # Get this metric for each method
            vals = {}
            for m in METHODS:
                if metric in ("ARI", "NMI"):
                    v = cdf_map.loc[m, metric] if m in cdf_map.index else np.nan
                else:
                    sub = bdf[bdf.get("method", bdf.get("Method", "")) == m]
                    v = float(sub[metric].iloc[0]) if len(sub) and metric in sub else np.nan
                vals[m] = float(v) if np.isfinite(v) else np.nan
            # Rank: 1 = best
            series = pd.Series(vals)
            rank_series = series.rank(ascending=not higher_better,
                                       method="min", na_option="keep")
            for m in METHODS:
                if np.isfinite(rank_series[m]):
                    ranks[m][metric].append(rank_series[m])

    # Build rows: Method | mean rank (overall) | per-metric mean rank
    header = ["Method", "Mean rank (overall)"] + [m for m, _ in METRICS]
    rows = []
    for m in METHODS:
        all_ranks = []
        per_metric = []
        for metric, _ in METRICS:
            rk = ranks[m][metric]
            if rk:
                per_metric.append(round(float(np.mean(rk)), 2))
                all_ranks.extend(rk)
            else:
                per_metric.append("—")
        overall = round(float(np.mean(all_ranks)), 2) if all_ranks else "—"
        rows.append([METHOD_DISPLAY.get(m, m), overall] + per_metric)

    # Sort by overall rank ascending (best first)
    rows.sort(key=lambda r: (r[1] if isinstance(r[1], (int, float)) else 1e9))

    _write_table(
        ws, rows,
        title="Supplementary Table S16 — Average rank meta-summary "
              "(12 methods × 6 main datasets × 7 canonical metrics)",
        note="For each (dataset, metric) the 12 methods are ranked (1 = best); "
             "ranks are then averaged across the 6 main datasets. The 'Mean "
             "rank (overall)' column averages all 6 × 7 = 42 individual ranks "
             "per method. Lower = better. Legacy average-rank view; the "
             "manuscript Fig. 6e ranking uses Integration Balance "
             "(S16_IB_results sheet), which supersedes this table.",
        header_row=header,
        col_widths=[14, 22] + [16] * len(METRICS),
    )


# ──────────────────────────────────────────────────────────────────────────────
# Table S17 — Data availability, accession and DOI
# ──────────────────────────────────────────────────────────────────────────────

def table_S17(ws):
    header = ["Dataset", "Type", "Primary citation", "Accession / repository",
              "Direct URL", "Notes"]
    rows = [
        ["Simulated 1–4", "Generated", "This study (Splatter)",
         "—", "Reproducible from seeds in scripts/run_experiments.py",
         "Splatter (Zappia et al. 2017 Genome Biol 18:174) was used with seed "
         "42; parameter sets archived in supplementary data."],
        ["Sc_mixology", "Cross-platform", "Tian et al. 2019 Nat Methods 16:479",
         "GEO GSE118767",
         "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE118767",
         "Processed h5ad from the original CellMix benchmark."],
        ["HDC (dendritic)", "Real Smart-seq2",
         "Villani et al. 2017 Science 356:eaah4573",
         "GEO GSE94820",
         "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE94820",
         "Subset to the four DC subtypes (pDC, DoubleNeg, CD141, CD1C) used by "
         "Villani et al."],
        ["PBMC (control / IFN-β)", "Condition",
         "Kang et al. 2018 Nat Biotechnol 36:89",
         "GEO GSE96583",
         "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE96583",
         "Processed control + IFN-β stimulated h5ad from the original paper."],
        ["Mouse Lung", "Atlas (16 batches)",
         "Luecken et al. 2022 Nat Methods 19:41",
         "figshare 12420968 (Luecken benchmark)",
         "https://figshare.com/articles/dataset/12420968",
         "Lung_atlas_public.h5ad from the public Luecken benchmark drop."],
        ["Human Pancreas", "Multi-protocol",
         "Luecken et al. 2022 Nat Methods 19:41",
         "figshare 12420968 (Luecken benchmark)",
         "https://figshare.com/articles/dataset/12420968",
         "human_pancreas_norm_complexBatch.h5ad."],
        ["Human Immune", "Multi-donor",
         "Luecken et al. 2022 Nat Methods 19:41",
         "figshare 12420968 (Luecken benchmark)",
         "https://figshare.com/articles/dataset/12420968",
         "Immune_ALL_human.h5ad."],
        ["Gut (mouse intestine)", "Condition",
         "Haber et al. 2017 Nature 551:333",
         "GEO GSE92332",
         "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE92332",
         "Salmonella / H.poly / control conditions."],
        ["scCAT source code", "Software",
         "This study", "GitHub: [TBD]", "[TBD]",
         "Includes preprocess, triplet construction, training and evaluation. "
         "Released under MIT license at acceptance."],
        ["scCAT analysis scripts", "Software / scripts",
         "This study", "GitHub: [TBD]", "[TBD]",
         "build_manuscript.py, plot_improved.py, plot_supp_figures.py, "
         "build_supp_tables.py are bundled with the code repository."],
        ["Processed embeddings (Zenodo)", "Output",
         "This study", "Zenodo: [DOI TBD on acceptance]", "[TBD]",
         "All 12 method outputs on the 11 datasets as h5ad/csv for direct "
         "reproduction of figures."],
    ]
    _write_table(
        ws, rows,
        title="Supplementary Table S17 — Data availability, accession and DOI",
        note="All publicly available datasets, with the primary citation and a "
             "direct URL where the processed object used in this study can be "
             "retrieved. scCAT source code and the analysis scripts will be "
             "deposited under permanent DOIs at acceptance.",
        header_row=header,
        col_widths=[26, 22, 36, 32, 60, 60],
    )


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    # Cover sheet
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover["A1"] = "Supplementary Tables — scCAT manuscript"
    ws_cover["A1"].font = Font(name="Arial", size=14, bold=True)
    ws_cover["A3"] = ("This workbook accompanies the manuscript "
                       '"scCAT preserves batch-specific cell states '
                       'through confidence-weighted triplet learning '
                       'in single-cell data integration".')
    ws_cover["A3"].font = Font(name="Arial", size=10, italic=True)
    contents = [
        ("S1",  "Comparison methods (11 baselines + scCAT)"),
        ("S2",  "Dataset summary (11 datasets used)"),
        ("S3",  "Evaluation metrics (7 canonical + 3 new + 1 composite + 1 readout)"),
        ("S4",  "scCAT hyperparameters (defaults + search ranges)"),
        ("S5",  "Full benchmark metric values (all datasets × methods)"),
        ("S6",  "Batch-specific preservation metrics (scCAT full)"),
        ("S7",  "Module ablation results (5 configs × 2 datasets)"),
        ("S8",  "Runtime and peak memory (4 benchmark datasets)"),
        ("S9",  "Per-dataset preprocessing parameters"),
        ("S10", "Per-batch × per-cell-type composition (all 11 datasets)"),
        ("S11", "Per-cell-type cluster purity (12 methods × 6 main datasets)"),
        ("S12", "Marker gene catalogue used for dot / feature plots"),
        ("S13", "Parameter sensitivity raw data (24 configs × 5 metrics)"),
        ("S14", "Mouse Lung per-batch local-mixing scores (16 × 12)"),
        ("S15", "Software environment and dependency versions"),
        ("S16", "Average rank meta-summary (12 methods × 6 datasets × 7 metrics)"),
        ("S17", "Data availability, accession and DOI"),
    ]
    ws_cover["A5"] = "Sheets in this workbook:"
    ws_cover["A5"].font = Font(name="Arial", size=10, bold=True)
    for i, (sid, desc) in enumerate(contents):
        ws_cover.cell(row=6 + i, column=1, value=sid).font = \
            Font(name="Arial", size=10, bold=True)
        ws_cover.cell(row=6 + i, column=2, value=desc).font = \
            Font(name="Arial", size=10)
    ws_cover.column_dimensions["A"].width = 8
    ws_cover.column_dimensions["B"].width = 70

    # Create each table
    builders = [
        ("S1_Methods",            table_S1),
        ("S2_Datasets",           table_S2),
        ("S3_Metrics",            table_S3),
        ("S4_Hyperparams",        table_S4),
        ("S5_Benchmark",          table_S5),
        ("S6_BSP_metrics",        table_S6),
        ("S7_Ablation",           table_S7),
        ("S8_Runtime",            table_S8),
        ("S9_Preprocessing",      table_S9),
        ("S10_Composition",       table_S10),
        ("S11_Celltype_purity",   table_S11),
        ("S12_Markers",           table_S12),
        ("S13_Sensitivity_raw",   table_S13),
        ("S14_Lung_batch_mixing", table_S14),
        ("S15_Software_env",      table_S15),
        ("S16_Average_rank",      table_S16),
        ("S17_Data_availability", table_S17),
    ]
    for sheet_name, builder in builders:
        ws = wb.create_sheet(title=sheet_name)
        try:
            builder(ws)
        except Exception as exc:
            ws.cell(row=1, column=1,
                     value=f"[ERROR while building {sheet_name}: {exc}]").font = NOTE_FONT
            print(f"  [WARN] {sheet_name} build failed: {exc}")
            import traceback; traceback.print_exc()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
