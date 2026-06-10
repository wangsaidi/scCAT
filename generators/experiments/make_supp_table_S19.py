"""
make_supp_table_S19.py — Append Supplementary Table S19 (HDC resolution-free
separability diagnostic) to the supplementary workbook, mirroring the existing
sheet style (cf. make_supp_table_S18.py / S18_scIB_standard).

Purpose: quantitatively back the §2.3 / Fig. 3c over-correction contrast.
The diagnostic is *resolution-free* (no clustering), so it shows that the HDC
over-correction is specific to the MNN family (Scanorama, fastMNN) and is not
an artefact of a particular k-means resolution. scCAT is in the well-behaved
group (1-NN same-type consistency 0.92) — it does NOT lead this ranking, which
is consistent with the main-text reading of HDC as evidence of absent
over-correction rather than a separation advantage over every competitor.

Source artifact (results/phase3/hdc_dc/, already computed + verified):
  separability_diagnostic.csv   method, loo_1nn, silhouette, ..., best_bs_ARI

Workbook:
  in : reproduce_tables/output/Supplementary_Tables_S1-S18.xlsx
  out: reproduce_tables/output/Supplementary_Tables_S1-S19.xlsx  (adds S19_HDC_separability, updates Cover)
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = Path(__file__).resolve().parent
HDC = HERE / "results" / "phase3" / "hdc_dc"
# Self-contained reproduction chain (see reproduce_tables/README.md): this
# script appends S19 to the S1-S18 workbook from make_supp_table_S18.py and
# writes S1-S19 under reproduce_tables/output/.  The shipped
# Publish_version/Supplementary_Tables_S1-S19.xlsx remains authoritative.
OUT_DIR = HERE.parent.parent / "reproduce_tables" / "output"
SRC_XLSX = OUT_DIR / "Supplementary_Tables_S1-S18.xlsx"
OUT_XLSX = OUT_DIR / "Supplementary_Tables_S1-S19.xlsx"

GREEN = "002D9E2D"
GREY = "00666666"
WHITE = "00FFFFFF"

DISP = {"INSCT_Unsupervised": "INSCT"}
def disp(m: str) -> str: return DISP.get(m, m)
SEMI = {"scANVI"}


def build_table() -> tuple[pd.DataFrame, dict]:
    d = pd.read_csv(HDC / "separability_diagnostic.csv")
    # integrity sanity: 1-NN consistency is a fraction
    assert d["loo_1nn"].between(0, 1).all(), "loo_1nn out of [0,1]"
    d = d.sort_values("loo_1nn", ascending=False).reset_index(drop=True)

    out = pd.DataFrame({
        "Method": [disp(m) for m in d["method"]],
        "1-NN same-type consistency": d["loo_1nn"].round(3),
        "Cell-type silhouette": d["silhouette"].round(3),
        "Best cDC1/cDC2 ARI": d["best_bs_ARI"].round(3),
        "Supervision": ["semi-supervised †" if disp(m) in SEMI else "unsupervised"
                        for m in d["method"]],
    })
    mnn = d[d["method"].isin(["Scanorama", "fastMNN"])].sort_values("method")
    others_min = float(d.loc[~d["method"].isin(["Scanorama", "fastMNN"]),
                             "loo_1nn"].min())
    stats = dict(
        sc_1nn=float(d.loc[d["method"] == "scCAT", "loo_1nn"].iloc[0]),
        sc_rank=int(d.index[d["method"] == "scCAT"][0]) + 1,
        n=len(d), others_min=others_min,
        scan_1nn=float(mnn.loc[mnn.method == "Scanorama", "loo_1nn"].iloc[0]),
        fmnn_1nn=float(mnn.loc[mnn.method == "fastMNN", "loo_1nn"].iloc[0]),
        scan_ari=float(mnn.loc[mnn.method == "Scanorama", "best_bs_ARI"].iloc[0]),
        fmnn_ari=float(mnn.loc[mnn.method == "fastMNN", "best_bs_ARI"].iloc[0]),
    )
    return out, stats


def write_sheet(out: pd.DataFrame, stats: dict) -> None:
    wb = openpyxl.load_workbook(SRC_XLSX)
    if "S19_HDC_separability" in wb.sheetnames:
        del wb["S19_HDC_separability"]
    ws = wb.create_sheet("S19_HDC_separability")
    ncol = out.shape[1]

    title = ("Supplementary Table S19 — HDC resolution-free separability "
             f"diagnostic ({stats['n']} methods)")
    note = (
        "Resolution-free check that the HDC over-correction in Fig. 3c is "
        "specific to the MNN family and is not an artefact of a particular "
        "clustering resolution. For each method we report (i) the "
        "leave-one-out 1-nearest-neighbour same-cell-type consistency "
        "(fraction of cells whose nearest neighbour in the integrated "
        "embedding is the same cell type — no clustering involved), (ii) the "
        "cell-type silhouette, and (iii) the best batch-specific cDC1/cDC2 "
        "separation (best ARI over a resolution sweep). The two MNN-family "
        "methods, Scanorama and fastMNN, are the clear outliers (1-NN "
        "consistency 0.68 and 0.74; best cDC1/cDC2 ARI 0.23 and 0.15), "
        "confirming that they intermix the batch-specific populations. Every "
        "other method scores at least 0.91 on 1-NN consistency (range "
        "0.915–0.995) and resolves the batch-specific subset far better; "
        "scCAT, at 0.92, sits at the lower end of this group but is well clear "
        "of the two over-correctors. scCAT is therefore "
        "in the well-behaved group rather than the over-correcting group; on "
        "this small two-batch dataset it does not lead the resolution-free "
        "separation ranking, consistent with the main-text reading of HDC as "
        "evidence of absent over-correction rather than a separation advantage "
        "over every competitor († scANVI is semi-supervised, uses labels). "
        "Numerical backing for the Fig. 3c / §2.3 over-correction contrast.")

    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws["A2"] = note
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color=GREY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws.row_dimensions[2].height = 150

    hdr_font = Font(name="Arial", bold=True, size=10, color=WHITE)
    hdr_fill = PatternFill("solid", fgColor=GREEN)
    for j, col in enumerate(out.columns, start=1):
        c = ws.cell(4, j, col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

    for i, (_, row) in enumerate(out.iterrows(), start=5):
        bold = (row["Method"] == "scCAT")
        for j, col in enumerate(out.columns, start=1):
            c = ws.cell(i, j, row[col])
            c.font = Font(name="Arial", size=10, bold=bold)
            c.alignment = Alignment(
                horizontal=("left" if j == 1 else "center"), vertical="center")

    widths = [12, 26, 20, 20, 18]
    for j, w in enumerate(widths[:ncol], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    cov = wb["Cover"]
    r = cov.max_row + 1
    cov.cell(r, 1, "S19").font = Font(name="Arial", size=10)
    cov.cell(r, 2, "HDC resolution-free separability diagnostic "
                   f"({stats['n']} methods)").font = Font(name="Arial", size=10)

    wb.save(OUT_XLSX)


def main() -> None:
    out, stats = build_table()
    write_sheet(out, stats)
    print("=== Supplementary Table S19 (HDC resolution-free separability) ===")
    print(out.to_string(index=False))
    print(f"\nscCAT 1-NN = {stats['sc_1nn']:.3f} (rank {stats['sc_rank']}/{stats['n']}); "
          f"MNN outliers Scanorama {stats['scan_1nn']:.3f} / fastMNN {stats['fmnn_1nn']:.3f}; "
          f"others_min = {stats['others_min']:.3f}")
    print("saved →", OUT_XLSX)


if __name__ == "__main__":
    main()
