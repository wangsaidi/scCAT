"""
make_supp_table_S18.py — Append Supplementary Table S18 (community-standard
scIB ranking robustness check, P0-2) to the supplementary workbook, mirroring
the existing sheet style (cf. S16_Average_rank).

Source artifacts (results/phase2/, all already computed + verified):
  scib_standard_secondary.csv            authoritative avg_rank + mean_scIB_overall + Friedman/Nemenyi stats
  scib_standard_per_method_dataset.csv   per (method, dataset) scIB_batch / bio / overall
  friedman_nemenyi_IBfull_secondary.csv  IB avg_rank (for Spearman rho vs IB)

Workbook:
  in : reproduce_tables/output/Supplementary_Tables_S1-S17.xlsx
  out: reproduce_tables/output/Supplementary_Tables_S1-S18.xlsx   (adds sheet S18_scIB_standard, updates Cover)

The mean scIB_overall recomputed here (same 9 datasets x 9 methods) is asserted
to match the authoritative secondary CSV, so the table cannot silently drift.
"""
from __future__ import annotations
import sys
from pathlib import Path
import numpy as np
import pandas as pd
from scipy.stats import spearmanr
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = Path(__file__).resolve().parent
PHASE2 = HERE / "results" / "phase2"
# Self-contained reproduction chain (see reproduce_tables/README.md): this
# script appends S18 to the S1-S17 workbook from build_supp_tables.py and writes
# S1-S18 under reproduce_tables/output/.  The shipped
# Publish_version/Supplementary_Tables_S1-S19.xlsx remains authoritative.
OUT_DIR = HERE.parent.parent / "reproduce_tables" / "output"
SRC_XLSX = OUT_DIR / "Supplementary_Tables_S1-S17.xlsx"
OUT_XLSX = OUT_DIR / "Supplementary_Tables_S1-S18.xlsx"

GREEN = "002D9E2D"           # scCAT green, matches S16 header fill
GREY = "00666666"
WHITE = "00FFFFFF"

DISP = {"INSCT_Unsupervised": "INSCT"}
def disp(m: str) -> str: return DISP.get(m, m)


# Nemenyi critical difference is a deterministic function of (k, N) at a fixed
# alpha: CD = q_alpha * sqrt(k*(k+1) / (6*N)), with q_alpha the Studentized-range
# quantile / sqrt(2) at alpha = 0.05. The upstream scib_standard_secondary.csv
# stored a CD with a rounding error (4.16); we recompute it here so the shipped
# table is exact (k = 9, N = 9 -> 4.00) and never depends on that stale field.
_NEMENYI_Q05 = {
    2: 1.960, 3: 2.343, 4: 2.569, 5: 2.728, 6: 2.850,
    7: 2.949, 8: 3.031, 9: 3.102, 10: 3.164,
}


def nemenyi_cd(k: int, N: int) -> float:
    """Nemenyi critical difference at alpha = 0.05 (deterministic in k, N)."""
    return _NEMENYI_Q05[k] * float(np.sqrt(k * (k + 1) / (6.0 * N)))


def build_table() -> tuple[pd.DataFrame, dict]:
    sec = pd.read_csv(PHASE2 / "scib_standard_secondary.csv")
    pmd = pd.read_csv(PHASE2 / "scib_standard_per_method_dataset.csv")
    ib = pd.read_csv(PHASE2 / "friedman_nemenyi_IBfull_secondary.csv")

    methods = list(sec["method"])
    # scCAT's 9 evaluation datasets (mirror the secondary-analysis restriction)
    sc_datasets = sorted(pmd.loc[pmd.method == "scCAT", "dataset"].unique())
    sub = pmd[pmd.dataset.isin(sc_datasets) & pmd.method.isin(methods)]

    agg = (sub.groupby("method")
              .agg(scIB_batch=("scIB_batch", "mean"),
                   scIB_bio=("scIB_bio", "mean"),
                   scIB_overall=("scIB_overall", "mean")))

    df = sec.merge(agg, on="method", how="left")
    # integrity gate: recomputed overall must match the authoritative CSV
    drift = (df["scIB_overall"] - df["mean_scIB_overall"]).abs().max()
    assert drift < 1e-6, f"overall mean drift {drift:.2e} — refusing to write"

    # Spearman rho between IB ranking and standard-scIB ranking (same 9 methods)
    merged = (ib[["method", "avg_rank"]].rename(columns={"avg_rank": "ib"})
                .merge(sec[["method", "avg_rank"]].rename(columns={"avg_rank": "scib"}),
                       on="method"))
    rho = float(spearmanr(merged["ib"], merged["scib"]).statistic)

    df = df.sort_values("avg_rank").reset_index(drop=True)
    out = pd.DataFrame({
        "Method": [disp(m) for m in df["method"]],
        "Avg. rank (lower = better)": df["avg_rank"].round(3),
        "Mean scIB overall": df["mean_scIB_overall"].round(4),
        "Mean scIB batch": df["scIB_batch"].round(4),
        "Mean scIB bio": df["scIB_bio"].round(4),
        "Supervision": ["semi-supervised †" if s else "unsupervised"
                        for s in df["semi_supervised"]],
    })
    stats = dict(chi2=float(sec["chi2"].iloc[0]), p=float(sec["p_value"].iloc[0]),
                 CD=nemenyi_cd(int(sec["k"].iloc[0]), int(sec["N"].iloc[0])),
                 k=int(sec["k"].iloc[0]),
                 N=int(sec["N"].iloc[0]), rho=rho,
                 sccat_rank=int(df.index[df["method"] == "scCAT"][0]) + 1,
                 top=df["method"].iloc[0])
    return out, stats


def write_sheet(out: pd.DataFrame, stats: dict) -> None:
    wb = openpyxl.load_workbook(SRC_XLSX)
    if "S18_scIB_standard" in wb.sheetnames:
        del wb["S18_scIB_standard"]
    ws = wb.create_sheet("S18_scIB_standard")
    ncol = out.shape[1]
    last = get_column_letter(ncol)

    title = ("Supplementary Table S18 — Community-standard scIB ranking "
             "robustness check (9 methods × 9 datasets)")
    note = (
        "Methods are re-ranked with the community-standard scIB aggregation "
        "(Luecken et al. 2022), scIB_overall = 0.4·s_batch + 0.6·s_bio, "
        "using the SAME component metrics as Integration Balance (IB) but the "
        "standard linear weighting instead of our geometric mean "
        "(s_batch = mean of iLISI, 1−kBET, ASW_batch_mixing; s_bio = mean of "
        "ARI, NMI, ASW_celltype, cLISI_purity). Unit of analysis = dataset "
        "(seeds averaged first); lower average rank = better. "
        f"Friedman χ² = {stats['chi2']:.1f}, p = {stats['p']:.1e}, "
        f"k = {stats['k']}, N = {stats['N']}; Nemenyi CD = {stats['CD']:.2f}. "
        f"scCAT ranks #{stats['sccat_rank']} overall and #1 among the "
        "unsupervised methods, behind only the semi-supervised scANVI "
        "(† uses ground-truth labels). The IB ranking (Table S16 / "
        f"Fig. 6e) and this standard-scIB ranking agree almost perfectly "
        f"(Spearman ρ = {stats['rho']:.2f}), i.e. the ordering is "
        "essentially invariant to the aggregation choice. Numerical source of "
        "Supplementary Figure S14.")

    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws["A2"] = note
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color=GREY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws.row_dimensions[2].height = 132

    hdr_font = Font(name="Arial", bold=True, size=10, color=WHITE)
    hdr_fill = PatternFill("solid", fgColor=GREEN)
    for j, col in enumerate(out.columns, start=1):
        c = ws.cell(4, j, col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, (_, row) in enumerate(out.iterrows(), start=5):
        for j, col in enumerate(out.columns, start=1):
            c = ws.cell(i, j, row[col])
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(
                horizontal=("left" if j == 1 else "center"), vertical="center")

    widths = [12, 22, 18, 18, 16, 18]
    for j, w in enumerate(widths[:ncol], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # update Cover index
    cov = wb["Cover"]
    r = cov.max_row + 1
    cov.cell(r, 1, "S18").font = Font(name="Arial", size=10)
    cov.cell(r, 2, "Community-standard scIB ranking robustness check "
                   "(9 methods × 9 datasets)").font = Font(name="Arial", size=10)

    wb.save(OUT_XLSX)


def main() -> None:
    out, stats = build_table()
    write_sheet(out, stats)
    print("=== Supplementary Table S18 (standard-scIB robustness) ===")
    print(out.to_string(index=False))
    print(f"\nFriedman chi2={stats['chi2']:.3f} p={stats['p']:.3e} "
          f"CD={stats['CD']:.3f} | Spearman rho(IB,scIB)={stats['rho']:.4f}")
    print(f"scCAT overall #{stats['sccat_rank']}; top method = {stats['top']}")
    print("saved →", OUT_XLSX)


if __name__ == "__main__":
    main()
